"""
宏耀考勤工资计算 - Vercel Serverless 入口
"""
import os
import sys
import re
import io
import calendar
from datetime import datetime

# 将项目根目录加入 path，以便导入 config 和 rules
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from flask import Flask, request, send_file, jsonify, Response
import openpyxl
from openpyxl import Workbook

from config import *
from rules.production import calculate as calc_production
from rules.production2 import calculate as calc_production2
from rules.mold import calculate as calc_mold
from rules.quality import calculate as calc_quality
from rules.tech import calculate as calc_tech
from rules.ouyang import calculate as calc_ouyang

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

RULE_FUNCTIONS = {
    'production':  (calc_production,  '生产部普工'),
    'production2': (calc_production2, '生产部2(计时)'),
    'mold':        (calc_mold,        '模房'),
    'quality':     (calc_quality,     '品质部'),
    'tech':        (calc_tech,        '技术部'),
    'ouyang':      (calc_ouyang,      '欧阳宇专属'),
}


def clean_name(name):
    return re.sub(r'[（(].+?[）)]', '', name).strip()


def get_employee_rule(name, department):
    cname = clean_name(name)
    if '离职' in name:
        return None, '离职跳过'
    if cname in SKIP_EMPLOYEES:
        return None, '免计算'
    if cname in SPECIAL_EMPLOYEES:
        rule_key = SPECIAL_EMPLOYEES[cname]
        func, display = RULE_FUNCTIONS[rule_key]
        return func, f'{display}(特殊)'
    dept = department.split('\n')[0].strip() if department else ''
    rule_key = DEPARTMENT_RULE_MAP.get(dept)
    if rule_key == 'skip':
        return None, '部门跳过'
    if rule_key and rule_key in RULE_FUNCTIONS:
        func, display = RULE_FUNCTIONS[rule_key]
        return func, display
    return None, '未识别部门'


def build_config(holidays=0):
    return {
        'night_snack_subsidy': NIGHT_SNACK_SUBSIDY,
        'full_attendance_bonus': FULL_ATTENDANCE_BONUS,
        'overtime_rate': OVERTIME_RATE,
        'production_base_salary': PRODUCTION_BASE_SALARY,
        'production_standard_days': PRODUCTION_STANDARD_DAYS,
        'production_standard_hours': PRODUCTION_STANDARD_HOURS,
        'production_weekday_ot_rate': PRODUCTION_WEEKDAY_OT_RATE,
        'production_weekend_ot_rate': PRODUCTION_WEEKEND_OT_RATE,
        'production2_hourly_rate': PRODUCTION2_HOURLY_RATE,
        'quality_monthly_rest_days': QUALITY_MONTHLY_REST_DAYS,
        'quality_base_salary': QUALITY_BASE_SALARY,
        'mold_hourly_rates': MOLD_HOURLY_RATES,
        'tech_sunday_rest': TECH_SUNDAY_REST,
        'tech_fixed_days': TECH_FIXED_DAYS,
        'tech_salary': TECH_SALARY,
        'high_temp_months': HIGH_TEMP_MONTHS,
        'ouyang_base': OUYANG_BASE,
        'ouyang_position': OUYANG_POSITION,
        'ouyang_subsidy': OUYANG_SUBSIDY,
        'holidays': holidays,
        'required_workdays': PRODUCTION_STANDARD_DAYS,
    }


def read_attendance_from_bytes(file_bytes, year, month):
    """从内存字节读取考勤表（不写临时文件）"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=False)
    if '打卡时间' in wb.sheetnames:
        ws = wb['打卡时间']
    else:
        ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    title = str(rows[0][0]) if rows[0][0] else ''
    date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})\s*至\s*(\d{4})-(\d{2})-(\d{2})', title)
    if date_match:
        start_day = int(date_match.group(3))
        end_day = int(date_match.group(6))
        total_days = end_day - start_day + 1
        first_day = start_day
    else:
        total_days = calendar.monthrange(year, month)[1]
        first_day = 1

    data_col_start = 6
    attendance = {}

    for row in rows[4:]:
        vals = list(row)
        name = vals[0]
        dept = vals[2]
        if not name or dept is None:
            continue
        name = str(name).strip()
        dept_str = str(dept).strip() if dept else ''

        records = []
        for day_idx in range(total_days):
            col_idx = data_col_start + day_idx
            day = first_day + day_idx
            punch_raw = vals[col_idx] if col_idx < len(vals) else None
            if punch_raw is not None and not isinstance(punch_raw, str):
                punch_raw = None
            punch_str = ''
            if punch_raw:
                punch_str = punch_raw.replace('\n', '  ').strip()
            weekday = calendar.weekday(year, month, day)
            records.append({
                'day': day,
                'punch_str': punch_str,
                'is_weekend': weekday >= 5,
            })

        attendance[name] = {
            'department': dept_str,
            'records': records,
        }

    return attendance, first_day, first_day + total_days - 1


def process_attendance(file_bytes, year, month, holidays=0):
    attendance, start_day, end_day = read_attendance_from_bytes(file_bytes, year, month)
    config = build_config(holidays)
    config['year'] = year
    config['month'] = month

    results = []
    all_anomalies = []
    skipped = []
    rule_groups = {}

    for name, info in attendance.items():
        records = info['records']
        department = info['department']
        calc_func, rule_name = get_employee_rule(name, department)

        if calc_func is None:
            skipped.append({'name': name, 'reason': rule_name, 'department': department})
            continue

        cname = clean_name(name)
        result = calc_func(cname, records, config, year, month)
        result['rule'] = rule_name
        result['department'] = department
        results.append(result)

        if result.get('anomalies'):
            all_anomalies.extend(result['anomalies'])

        rule_groups.setdefault(rule_name, []).append(cname)

    return {
        'results': results,
        'anomalies': all_anomalies,
        'skipped': skipped,
        'rule_groups': rule_groups,
        'date_range': f"{year}年{month}月{start_day}日~{end_day}日",
        'total_employees': len(attendance),
    }


def build_excel_bytes(data, year, month):
    """生成 Excel 到内存并返回 bytes"""
    wb = Workbook()
    ws = wb.active
    ws.title = "工资计算结果"
    results = data['results']

    all_keys = set()
    for r in results:
        all_keys.update(k for k in r.keys() if k != 'anomalies')

    priority_keys = [
        'name', 'department', 'rule',
        'required_days', 'actual_days', 'workday_days', 'weekend_days',
        'total_hours', 'hourly_rate',
        'base_salary', 'position_salary',
        'weekday_ot_hours', 'weekday_ot_salary',
        'overtime_hours', 'ot_hourly_rate', 'overtime_salary',
        'weekend_hours', 'weekend_salary',
        'night_shift_days', 'night_snack_subsidy',
        'full_attendance_bonus', 'high_temp_subsidy',
        'subsidy', 'holidays', 'holiday_salary',
        'total_salary',
    ]
    headers = [k for k in priority_keys if k in all_keys]
    headers += sorted(all_keys - set(headers))

    cn = {
        'name': '姓名', 'department': '部门', 'rule': '规则类型',
        'required_days': '应出勤天数', 'actual_days': '实际出勤天数',
        'workday_days': '工作日出勤', 'weekend_days': '周末出勤',
        'total_hours': '总工时', 'hourly_rate': '时薪',
        'base_salary': '基本工资', 'position_salary': '岗位工资',
        'weekday_ot_hours': '工作日加班工时', 'weekday_ot_salary': '工作日加班工资',
        'overtime_hours': '加班工时', 'ot_hourly_rate': '加班时薪',
        'overtime_salary': '加班工资',
        'weekend_hours': '周末工时', 'weekend_salary': '周末工资',
        'night_shift_days': '夜班天数', 'night_snack_subsidy': '夜宵补贴',
        'full_attendance_bonus': '全勤奖', 'high_temp_subsidy': '高温补贴',
        'subsidy': '补贴', 'holidays': '法定节假日天数',
        'holiday_salary': '节假日工资', 'total_salary': '应发工资合计',
    }

    ws.append([cn.get(h, h) for h in headers])
    for r in results:
        ws.append([r.get(h, '') for h in headers])

    if data['anomalies']:
        ws2 = wb.create_sheet("异常报告")
        ws2.append(['姓名', '日期', '异常类型', '打卡内容'])
        for a in data['anomalies']:
            ws2.append([a['name'], a['date'], a['type'], a['punch']])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------- HTML 页面（内嵌，避免 Vercel 模板路径问题）----------

INDEX_HTML = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>宏耀考勤工资计算</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,"Microsoft YaHei",sans-serif;background:#f0f2f5;color:#333}
.container{max-width:1200px;margin:0 auto;padding:20px}
.header{text-align:center;padding:30px 0 20px}
.header h1{font-size:24px;color:#1a1a1a}
.header p{color:#666;margin-top:8px}
.upload-card{background:#fff;border-radius:12px;padding:30px;box-shadow:0 2px 8px rgba(0,0,0,.08);margin-bottom:20px}
.form-row{display:flex;gap:16px;align-items:flex-end;flex-wrap:wrap}
.form-group{display:flex;flex-direction:column;gap:6px}
.form-group label{font-size:14px;color:#666;font-weight:500}
.form-group select,.form-group input[type="number"]{padding:8px 12px;border:1px solid #d9d9d9;border-radius:6px;font-size:14px;background:#fff;min-width:100px}
.upload-area{border:2px dashed #d9d9d9;border-radius:8px;padding:40px;text-align:center;cursor:pointer;transition:all .3s;margin:20px 0}
.upload-area:hover,.upload-area.dragover{border-color:#1677ff;background:#f0f5ff}
.upload-area .icon{font-size:40px;color:#1677ff}
.upload-area p{color:#666;margin-top:8px}
.upload-area .filename{color:#1677ff;font-weight:500;margin-top:8px}
.btn{padding:10px 28px;border:none;border-radius:6px;font-size:15px;cursor:pointer;transition:all .3s;font-weight:500;text-decoration:none;display:inline-block}
.btn-primary{background:#1677ff;color:#fff}
.btn-primary:hover{background:#4096ff}
.btn-primary:disabled{background:#d9d9d9;cursor:not-allowed}
.btn-success{background:#52c41a;color:#fff}
.btn-success:hover{background:#73d13d}
.loading{display:none;text-align:center;padding:40px}
.loading .spinner{width:40px;height:40px;border:4px solid #f0f0f0;border-top:4px solid #1677ff;border-radius:50%;animation:spin 1s linear infinite;margin:0 auto 16px}
@keyframes spin{to{transform:rotate(360deg)}}
.result-section{display:none}
.stats-row{display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap}
.stat-card{background:#fff;border-radius:8px;padding:16px 20px;flex:1;min-width:140px;box-shadow:0 1px 4px rgba(0,0,0,.06)}
.stat-card .label{font-size:13px;color:#999}
.stat-card .value{font-size:24px;font-weight:600;margin-top:4px}
.stat-card .value.blue{color:#1677ff}
.stat-card .value.green{color:#52c41a}
.stat-card .value.orange{color:#fa8c16}
.group-tags{margin-bottom:20px}
.group-tag{display:inline-block;padding:4px 12px;margin:4px;background:#f0f5ff;color:#1677ff;border-radius:4px;font-size:13px}
.table-card{background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.08);margin-bottom:20px}
.tabs{display:flex;gap:0;border-bottom:1px solid #f0f0f0}
.tab{padding:10px 20px;cursor:pointer;border-bottom:2px solid transparent;font-size:14px;color:#666;transition:all .3s}
.tab.active{color:#1677ff;border-bottom-color:#1677ff;font-weight:500}
.tab-content{display:none}
.tab-content.active{display:block}
.table-wrapper{overflow-x:auto}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#fafafa;padding:10px 12px;text-align:left;font-weight:500;color:#666;white-space:nowrap}
td{padding:10px 12px;border-top:1px solid #f0f0f0;white-space:nowrap}
tr:hover td{background:#f5f5f5}
.salary-cell{font-weight:600;color:#1677ff}
.download-bar{position:sticky;bottom:0;background:#fff;padding:16px 20px;border-top:1px solid #f0f0f0;text-align:center;box-shadow:0 -2px 8px rgba(0,0,0,.06)}
</style>
</head>
<body>
<div class="container">
<div class="header">
<h1>宏耀考勤工资计算系统</h1>
<p>上传钉钉考勤报表，自动计算每位员工工资</p>
</div>

<div class="upload-card">
<form id="uploadForm">
<div class="form-row">
<div class="form-group"><label>年份</label><select name="year" id="year">YEAR_OPTIONS</select></div>
<div class="form-group"><label>月份</label><select name="month" id="month">MONTH_OPTIONS</select></div>
<div class="form-group"><label>法定节假日天数</label><input type="number" name="holidays" id="holidays" value="0" min="0" max="10"></div>
<div class="form-group"><button type="submit" class="btn btn-primary" id="calcBtn" disabled>开始计算</button></div>
</div>
<div class="upload-area" id="dropZone">
<div class="icon">📁</div>
<p>点击上传或拖拽考勤表到此处</p>
<p style="font-size:12px;color:#999">支持 .xlsx 格式（钉钉导出的考勤报表）</p>
<div class="filename" id="fileName"></div>
<input type="file" id="fileInput" name="file" accept=".xlsx" style="display:none">
</div>
</form>
</div>

<div class="loading" id="loading"><div class="spinner"></div><p>正在计算工资，请稍候...</p></div>

<div class="result-section" id="resultSection">
<div class="stats-row">
<div class="stat-card"><div class="label">日期范围</div><div class="value" id="statRange" style="font-size:16px"></div></div>
<div class="stat-card"><div class="label">总人数</div><div class="value blue" id="statTotal"></div></div>
<div class="stat-card"><div class="label">已计算</div><div class="value green" id="statCalc"></div></div>
<div class="stat-card"><div class="label">跳过</div><div class="value" id="statSkip"></div></div>
<div class="stat-card"><div class="label">异常记录</div><div class="value orange" id="statAnomaly"></div></div>
</div>
<div class="group-tags" id="groupTags"></div>
<div class="table-card">
<div class="tabs" id="tabs">
<div class="tab active" data-tab="salary">工资明细</div>
<div class="tab" data-tab="anomaly">异常报告</div>
<div class="tab" data-tab="skipped">跳过人员</div>
</div>
<div class="tab-content active" id="tab-salary"><div class="table-wrapper"><table id="salaryTable"><thead></thead><tbody></tbody></table></div></div>
<div class="tab-content" id="tab-anomaly"><div class="table-wrapper"><table id="anomalyTable"><thead></thead><tbody></tbody></table></div></div>
<div class="tab-content" id="tab-skipped"><div class="table-wrapper"><table id="skippedTable"><thead></thead><tbody></tbody></table></div></div>
</div>
<div class="download-bar"><a id="downloadBtn" class="btn btn-success">下载 Excel 结果</a></div>
</div>
</div>

<script>
const dropZone=document.getElementById('dropZone'),fileInput=document.getElementById('fileInput'),calcBtn=document.getElementById('calcBtn');
let selectedFile=null;
dropZone.addEventListener('click',()=>fileInput.click());
fileInput.addEventListener('change',e=>{if(e.target.files[0])selectFile(e.target.files[0])});
dropZone.addEventListener('dragover',e=>{e.preventDefault();dropZone.classList.add('dragover')});
dropZone.addEventListener('dragleave',()=>dropZone.classList.remove('dragover'));
dropZone.addEventListener('drop',e=>{e.preventDefault();dropZone.classList.remove('dragover');if(e.dataTransfer.files[0])selectFile(e.dataTransfer.files[0])});
function selectFile(f){if(!f.name.endsWith('.xlsx')){alert('请上传 .xlsx 格式文件');return}selectedFile=f;document.getElementById('fileName').textContent=f.name;calcBtn.disabled=false}
document.getElementById('uploadForm').addEventListener('submit',async e=>{
e.preventDefault();if(!selectedFile)return;
const fd=new FormData();fd.append('file',selectedFile);fd.append('year',document.getElementById('year').value);fd.append('month',document.getElementById('month').value);fd.append('holidays',document.getElementById('holidays').value);
document.getElementById('loading').style.display='block';document.getElementById('resultSection').style.display='none';calcBtn.disabled=true;
try{const r=await fetch('/calculate',{method:'POST',body:fd});const d=await r.json();if(d.error){alert(d.error);return}renderResults(d)}catch(e){alert('请求失败: '+e.message)}finally{document.getElementById('loading').style.display='none';calcBtn.disabled=false}
});
document.getElementById('tabs').addEventListener('click',e=>{if(!e.target.classList.contains('tab'))return;document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));document.querySelectorAll('.tab-content').forEach(t=>t.classList.remove('active'));e.target.classList.add('active');document.getElementById('tab-'+e.target.dataset.tab).classList.add('active')});
const fn={required_days:'应出勤',actual_days:'实际出勤',workday_days:'工作日出勤',weekend_days:'周末出勤',total_hours:'总工时',hourly_rate:'时薪',base_salary:'基本工资',position_salary:'岗位工资',weekday_ot_hours:'工作日加班H',weekday_ot_salary:'工作日加班工资',overtime_hours:'加班H',ot_hourly_rate:'加班时薪',overtime_salary:'加班工资',weekend_hours:'周末工时',weekend_salary:'周末工资',night_shift_days:'夜班天数',night_snack_subsidy:'夜宵补贴',full_attendance_bonus:'全勤奖',high_temp_subsidy:'高温补贴',subsidy:'补贴',holidays:'节假日天数',holiday_salary:'节假日工资',total_salary:'应发合计'};
function renderResults(d){
document.getElementById('statRange').textContent=d.date_range;document.getElementById('statTotal').textContent=d.total_employees;document.getElementById('statCalc').textContent=d.calculated;document.getElementById('statSkip').textContent=d.skipped_count;document.getElementById('statAnomaly').textContent=d.anomaly_count;
document.getElementById('groupTags').innerHTML=Object.entries(d.rule_groups).map(([r,n])=>`<span class="group-tag">${r} (${n.length}人)</span>`).join('');
const dk=new Set();d.results.forEach(r=>Object.keys(r.details).forEach(k=>dk.add(k)));
const ko=['required_days','actual_days','workday_days','weekend_days','total_hours','hourly_rate','base_salary','position_salary','weekday_ot_hours','weekday_ot_salary','overtime_hours','ot_hourly_rate','overtime_salary','weekend_hours','weekend_salary','night_shift_days','night_snack_subsidy','full_attendance_bonus','high_temp_subsidy','subsidy','holidays','holiday_salary','total_salary'];
const ok=ko.filter(k=>dk.has(k));ok.push(...[...dk].filter(k=>!ko.includes(k)).sort());
let th='<tr><th>姓名</th><th>部门</th><th>规则</th>';ok.forEach(k=>{th+=`<th>${fn[k]||k}</th>`});th+='</tr>';
let tb='';d.results.forEach(r=>{tb+=`<tr><td>${r.name}</td><td>${r.department}</td><td>${r.rule}</td>`;ok.forEach(k=>{let v=r.details[k];if(v===null||v===undefined||v==='')v='-';else if(typeof v==='number')v=Number.isInteger(v)?v:v.toFixed(2);const c=k==='total_salary'?' class="salary-cell"':'';tb+=`<td${c}>${v}</td>`});tb+='</tr>'});
document.querySelector('#salaryTable thead').innerHTML=th;document.querySelector('#salaryTable tbody').innerHTML=tb;
document.querySelector('#anomalyTable thead').innerHTML='<tr><th>姓名</th><th>日期</th><th>类型</th><th>打卡</th></tr>';
document.querySelector('#anomalyTable tbody').innerHTML=`<tr><td colspan="4" style="text-align:center;color:#999;padding:20px">共 ${d.anomaly_count} 条异常，请下载 Excel 查看详情</td></tr>`;
document.querySelector('#skippedTable thead').innerHTML='<tr><th>姓名</th><th>部门</th><th>原因</th></tr>';
let sb='';d.skipped.forEach(s=>{sb+=`<tr><td>${s.name}</td><td>${s.department}</td><td>${s.reason}</td></tr>`});
document.querySelector('#skippedTable tbody').innerHTML=sb||'<tr><td colspan="3" style="text-align:center;color:#999">无</td></tr>';
document.getElementById('downloadBtn').href='/download/'+encodeURIComponent(d.download_file);
document.getElementById('resultSection').style.display='block';document.getElementById('resultSection').scrollIntoView({behavior:'smooth'});
}
</script>
</body>
</html>"""

# 全局缓存最近一次计算的 Excel（Vercel 无持久文件系统）
_last_excel_cache = {}


@app.route('/')
def index():
    now = datetime.now()
    year_opts = ''.join(
        f'<option value="{y}" {"selected" if y==now.year else ""}>{y}年</option>'
        for y in range(now.year - 1, now.year + 2)
    )
    month_opts = ''.join(
        f'<option value="{m}" {"selected" if m==now.month else ""}>{m}月</option>'
        for m in range(1, 13)
    )
    html = INDEX_HTML.replace('YEAR_OPTIONS', year_opts).replace('MONTH_OPTIONS', month_opts)
    return Response(html, content_type='text/html; charset=utf-8')


@app.route('/calculate', methods=['POST'])
def calculate():
    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        return jsonify({'error': '请上传 .xlsx 格式的考勤表'}), 400

    year = int(request.form.get('year', datetime.now().year))
    month = int(request.form.get('month', datetime.now().month))
    holidays = int(request.form.get('holidays', 0))

    try:
        file_bytes = file.read()
        data = process_attendance(file_bytes, year, month, holidays)
        excel_buf = build_excel_bytes(data, year, month)

        # 缓存到内存供下载
        filename = f"工资计算_{year}年{month}月.xlsx"
        _last_excel_cache['filename'] = filename
        _last_excel_cache['data'] = excel_buf.getvalue()

        display_results = []
        for r in data['results']:
            display_results.append({
                'name': r.get('name', ''),
                'department': r.get('department', ''),
                'rule': r.get('rule', ''),
                'total_salary': round(r.get('total_salary', 0) or 0, 2),
                'details': {k: v for k, v in r.items()
                           if k not in ('name', 'department', 'rule', 'anomalies')},
            })

        return jsonify({
            'success': True,
            'date_range': data['date_range'],
            'total_employees': data['total_employees'],
            'calculated': len(data['results']),
            'skipped_count': len(data['skipped']),
            'anomaly_count': len(data['anomalies']),
            'results': display_results,
            'skipped': data['skipped'],
            'rule_groups': data['rule_groups'],
            'download_file': filename,
        })
    except Exception as e:
        return jsonify({'error': f'计算出错: {str(e)}'}), 500


@app.route('/download/<filename>')
def download(filename):
    if _last_excel_cache.get('filename') == filename and _last_excel_cache.get('data'):
        return send_file(
            io.BytesIO(_last_excel_cache['data']),
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    return '文件不存在或已过期，请重新计算', 404
