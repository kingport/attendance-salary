"""
宏耀考勤工资计算工具 - 主入口
用法: python main.py <考勤表.xlsx> [--year YYYY] [--month MM] [--holidays N]
"""
import sys
import os
import re
import argparse
import calendar
from datetime import datetime

import openpyxl
from openpyxl import Workbook

from config import *
from rules.production import calculate as calc_production
from rules.production2 import calculate as calc_production2
from rules.mold import calculate as calc_mold
from rules.quality import calculate as calc_quality
from rules.tech import calculate as calc_tech
from rules.ouyang import calculate as calc_ouyang

# 规则名 → 计算函数 + 显示名
RULE_FUNCTIONS = {
    'production':  (calc_production,  '生产部普工'),
    'production2': (calc_production2, '生产部2(计时)'),
    'mold':        (calc_mold,        '模房'),
    'quality':     (calc_quality,     '品质部'),
    'tech':        (calc_tech,        '技术部'),
    'ouyang':      (calc_ouyang,      '欧阳宇专属'),
}


def clean_name(name):
    """去除姓名中的（离职）等后缀"""
    return re.sub(r'[（(].+?[）)]', '', name).strip()


def get_employee_rule(name, department):
    """
    根据员工姓名和部门确定计算规则
    优先级：离职标记 > 免计算名单 > 特殊人员 > 部门映射
    """
    cname = clean_name(name)

    # 1. 名字里带"离职"直接跳过
    if '离职' in name:
        return None, '离职跳过'

    # 2. 免计算名单
    if cname in SKIP_EMPLOYEES:
        return None, '免计算'

    # 3. 特殊人员（指定规则，不按部门）
    if cname in SPECIAL_EMPLOYEES:
        rule_key = SPECIAL_EMPLOYEES[cname]
        func, display = RULE_FUNCTIONS[rule_key]
        return func, f'{display}(特殊)'

    # 4. 按部门映射
    # 部门字段可能包含多个部门（换行分隔），取第一个有效的
    dept = department.split('\n')[0].strip() if department else ''

    rule_key = DEPARTMENT_RULE_MAP.get(dept)
    if rule_key == 'skip':
        return None, '部门跳过'
    if rule_key and rule_key in RULE_FUNCTIONS:
        func, display = RULE_FUNCTIONS[rule_key]
        return func, display

    return None, '未识别部门'


def build_config(holidays=0):
    """组装配置 dict 供规则函数使用"""
    return {
        'night_snack_subsidy': NIGHT_SNACK_SUBSIDY,
        'full_attendance_bonus': FULL_ATTENDANCE_BONUS,
        'overtime_rate': OVERTIME_RATE,
        'production_base_salary': PRODUCTION_BASE_SALARY,
        'production_standard_days': PRODUCTION_STANDARD_DAYS,
        'production_standard_hours': PRODUCTION_STANDARD_HOURS,
        'production_weekday_ot_rate': PRODUCTION_WEEKDAY_OT_RATE,
        'production_weekend_ot_rate': PRODUCTION_WEEKEND_OT_RATE,
        'production2_hourly_rate': PRODUCTION2_HOURLY_RATE,
        'quality_monthly_rest_days': QUALITY_MONTHLY_REST_DAYS,
        'quality_base_salary': QUALITY_BASE_SALARY,
        'mold_hourly_rates': MOLD_HOURLY_RATES,
        'tech_sunday_rest': TECH_SUNDAY_REST,
        'tech_fixed_days': TECH_FIXED_DAYS,
        'tech_salary': TECH_SALARY,
        'high_temp_months': HIGH_TEMP_MONTHS,
        'ouyang_base': OUYANG_BASE,
        'ouyang_position': OUYANG_POSITION,
        'ouyang_subsidy': OUYANG_SUBSIDY,
        'holidays': holidays,
        'required_workdays': PRODUCTION_STANDARD_DAYS,
    }


def read_attendance(filepath, year, month):
    """
    读取钉钉考勤报表 Excel 的 "打卡时间" sheet

    返回: dict {
        员工姓名: {
            'department': str,
            'records': [{'day': int, 'punch_str': str, 'is_weekend': bool}, ...]
        }
    }
    """
    wb = openpyxl.load_workbook(filepath, read_only=False)
    if '打卡时间' in wb.sheetnames:
        ws = wb['打卡时间']
    else:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # 从标题行解析日期范围
    title = str(rows[0][0]) if rows[0][0] else ''
    date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})\s*至\s*(\d{4})-(\d{2})-(\d{2})', title)
    if date_match:
        start_day = int(date_match.group(3))
        end_day = int(date_match.group(6))
        total_days = end_day - start_day + 1
        first_day = start_day
    else:
        total_days = calendar.monthrange(year, month)[1]
        first_day = 1

    print(f"日期范围: {year}年{month}月{first_day}日 ~ {first_day + total_days - 1}日 ({total_days}天)")

    data_col_start = 6
    attendance = {}

    for row in rows[4:]:
        vals = list(row)
        name = vals[0]
        dept = vals[2]

        if not name or dept is None:
            continue

        name = str(name).strip()
        dept_str = str(dept).strip() if dept else ''

        records = []
        for day_idx in range(total_days):
            col_idx = data_col_start + day_idx
            day = first_day + day_idx

            punch_raw = vals[col_idx] if col_idx < len(vals) else None
            if punch_raw is not None and not isinstance(punch_raw, str):
                punch_raw = None

            punch_str = ''
            if punch_raw:
                punch_str = punch_raw.replace('\n', '  ').strip()

            weekday = calendar.weekday(year, month, day)

            records.append({
                'day': day,
                'punch_str': punch_str,
                'is_weekend': weekday >= 5,
            })

        attendance[name] = {
            'department': dept_str,
            'records': records,
        }

    return attendance


def run(filepath, year, month, holidays=0):
    """主计算流程"""
    print(f"=== 宏耀考勤工资计算 {year}年{month}月 ===")
    if holidays:
        print(f"法定节假日: {holidays}天")
    print()

    attendance = read_attendance(filepath, year, month)
    print(f"共 {len(attendance)} 名员工\n")

    config = build_config(holidays)
    config['year'] = year
    config['month'] = month

    results = []
    all_anomalies = []
    skipped = []
    unassigned = []

    # 分类汇总
    rule_groups = {}

    for name, info in attendance.items():
        records = info['records']
        department = info['department']
        calc_func, rule_name = get_employee_rule(name, department)

        if calc_func is None:
            skipped.append(f"{name}({rule_name})")
            continue

        cname = clean_name(name)
        result = calc_func(cname, records, config, year, month)
        result['rule'] = rule_name
        result['department'] = department
        results.append(result)

        if result.get('anomalies'):
            all_anomalies.extend(result['anomalies'])

        # 分组统计
        rule_groups.setdefault(rule_name, []).append(cname)

    # 打印分类结果
    print("--- 人员分类 ---")
    for rule, names in rule_groups.items():
        print(f"  [{rule}] ({len(names)}人): {', '.join(names)}")
    if skipped:
        print(f"  [跳过] ({len(skipped)}人): {', '.join(skipped)}")
    if unassigned:
        print(f"  [未识别] ({len(unassigned)}人): {', '.join(unassigned)}")
    print()

    # 输出结果
    output_dir = os.path.join(os.path.dirname(__file__), 'output')
    os.makedirs(output_dir, exist_ok=True)

    output_file = os.path.join(output_dir, f"工资计算_{year}年{month}月.xlsx")
    write_results(results, output_file)
    print(f"工资结果已保存到: {output_file}")

    if all_anomalies:
        anomaly_file = os.path.join(output_dir, f"异常报告_{year}年{month}月.xlsx")
        write_anomalies(all_anomalies, anomaly_file)
        print(f"异常报告已保存到: {anomaly_file}")

    print(f"\n--- 汇总 ---")
    print(f"已计算: {len(results)} 人")
    print(f"跳过: {len(skipped)} 人")
    if all_anomalies:
        print(f"异常记录: {len(all_anomalies)} 条")


def write_results(results, filepath):
    """将计算结果写入 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "工资计算结果"

    all_keys = set()
    for r in results:
        all_keys.update(k for k in r.keys() if k != 'anomalies')

    priority_keys = [
        'name', 'department', 'rule',
        'required_days', 'actual_days', 'workday_days', 'weekend_days',
        'total_hours', 'hourly_rate',
        'base_salary', 'position_salary',
        'weekday_ot_hours', 'weekday_ot_salary',
        'overtime_hours', 'ot_hourly_rate', 'overtime_salary',
        'weekend_hours', 'weekend_salary',
        'night_shift_days', 'night_snack_subsidy',
        'full_attendance_bonus', 'high_temp_subsidy',
        'subsidy', 'holidays', 'holiday_salary',
        'total_salary',
    ]
    headers = [k for k in priority_keys if k in all_keys]
    headers += sorted(all_keys - set(headers))

    cn_headers = {
        'name': '姓名', 'department': '部门', 'rule': '规则类型',
        'required_days': '应出勤天数', 'actual_days': '实际出勤天数',
        'workday_days': '工作日出勤', 'weekend_days': '周末出勤',
        'total_hours': '总工时', 'hourly_rate': '时薪',
        'base_salary': '基本工资', 'position_salary': '岗位工资',
        'weekday_ot_hours': '工作日加班工时', 'weekday_ot_salary': '工作日加班工资',
        'overtime_hours': '加班工时', 'ot_hourly_rate': '加班时薪',
        'overtime_salary': '加班工资',
        'weekend_hours': '周末工时', 'weekend_salary': '周末工资',
        'night_shift_days': '夜班天数', 'night_snack_subsidy': '夜宵补贴',
        'full_attendance_bonus': '全勤奖', 'high_temp_subsidy': '高温补贴',
        'subsidy': '补贴', 'holidays': '法定节假日天数',
        'holiday_salary': '节假日工资', 'total_salary': '应发工资合计',
    }

    ws.append([cn_headers.get(h, h) for h in headers])
    for r in results:
        row = [r.get(h, '') for h in headers]
        ws.append(row)

    wb.save(filepath)


def write_anomalies(anomalies, filepath):
    """将异常记录写入 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "异常报告"
    ws.append(['姓名', '日期', '异常类型', '打卡内容'])
    for a in anomalies:
        ws.append([a['name'], a['date'], a['type'], a['punch']])
    wb.save(filepath)


def main():
    parser = argparse.ArgumentParser(description='宏耀考勤工资计算工具')
    parser.add_argument('input', help='考勤表 Excel 文件路径')
    parser.add_argument('--year', type=int, default=datetime.now().year, help='年份')
    parser.add_argument('--month', type=int, default=datetime.now().month, help='月份')
    parser.add_argument('--holidays', type=int, default=0, help='当月法定节假日天数')
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"错误: 文件不存在 - {args.input}")
        sys.exit(1)

    run(args.input, args.year, args.month, args.holidays)


if __name__ == '__main__':
    main()
