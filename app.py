"""
宏耀考勤工资计算 - Web 应用
启动: python3 app.py
访问: http://localhost:5001
"""
import os
import re
import calendar
from datetime import datetime

from flask import Flask, render_template, request, jsonify
import openpyxl

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB

# 生产部排除人员
PRODUCTION_EXCLUDE = ['王琴', '欧阳宇', '李乐平', '熊其享']

# 模房排除人员（有特殊规则）
MOLD_EXCLUDE = ['张红亮', '张翱']

# 技术部：从其他部门走技术部规则的特殊人员
TECH_SPECIAL = ['李乐平', '熊其享', '张红亮', '耿红志']

# 技术部：应出勤 = 当月天数 - 周日天数 的人员
TECH_SUNDAY_REST = ['耿红志', '张红亮', '黎钦德', '李乐平', '熊其享']

# 技术部：固定底薪 + 岗位工资
TECH_SALARY = {
    '耿红志': {'base': 3730, 'position': 4270},
    '张红亮': {'base': 3730, 'position': 8770},
    '黎钦德': {'base': 3730, 'position': 7970},
    '刘高伟': {'base': 4020, 'position': 3980},
    '陶发志': {'base': 4020, 'position': 5180},
    '王德全': {'base': 4020, 'position': 5480},
    '熊庆':   {'base': 2730, 'position': 4270},
    '李乐平': {'base': 2730, 'position': 3270},
    '熊其享': {'base': 2730, 'position': 2270},
}


def clean_name(name):
    """去除姓名中的（离职）等后缀"""
    return re.sub(r'[（(].+?[）)]', '', name).strip()


def parse_time(time_str):
    """解析 HH:MM 时间字符串"""
    try:
        return datetime.strptime(time_str.strip(), "%H:%M")
    except (ValueError, AttributeError):
        return None


def parse_punch_records(punch_str):
    """解析打卡记录，保持原始顺序（夜班先晚后早）"""
    if not punch_str or not str(punch_str).strip():
        return []
    times = []
    for part in str(punch_str).strip().split():
        t = parse_time(part)
        if t:
            times.append(t)
    return times


def round_half_down(hours):
    """工时取整：0.5H 阶梯向下取整"""
    if hours < 0:
        return 0
    return int(hours * 2) / 2


def detect_shift(times):
    """
    判断班次：
    - 首次打卡在 19:00~20:00 之间 → 夜班
    - 否则 → 白班
    """
    if not times:
        return 'unknown'
    t19 = parse_time("19:00")
    t20 = parse_time("20:00")
    return 'night' if t19 <= times[0] <= t20 else 'day'


def calc_work_hours(times, shift):
    """
    计算单日工时
    白班：最后打卡 - 最早打卡
    夜班：固定4h（当晚20:00~24:00）+ 最后打卡时间（零点起算）
          例: 19:50~09:00 → 4 + 9 = 13h
    """
    if len(times) < 2:
        return 0

    if shift == 'night':
        # 夜班：当晚固定4h + 次日零点到最后打卡的时间
        last_punch = times[-1]
        morning_hours = last_punch.hour + last_punch.minute / 60
        return round_half_down(4 + morning_hours)
    else:
        # 白班：末次 - 首次
        total = (times[-1] - times[0]).total_seconds() / 3600
        return round_half_down(total)


def calc_production_employee(name, records, year, month):
    """
    计算生产部单个员工的出勤和工资
    规则：
    - 白班/夜班：首次打卡在19:00-20:00之间为夜班
    - 夜班工时 = 4h + 最后打卡时间（零点起算）
    - 夜班补贴：15元/天
    - 工作日出勤超8小时为加班，加班费17.9元/h
    - 周末出勤按23.86元/h
    - 基本工资 = 2100/22 * 工作日出勤天数
    """
    workday_days = 0
    weekend_days = 0
    weekday_ot_hours = 0
    weekend_hours = 0
    day_shift_count = 0
    night_shift_count = 0

    for rec in records:
        day = rec['day']
        punch_str = rec.get('punch_str', '')
        times = parse_punch_records(punch_str)

        if len(times) < 2:
            continue

        shift = detect_shift(times)
        total_hours = calc_work_hours(times, shift)

        if shift == 'day':
            day_shift_count += 1
        else:
            night_shift_count += 1

        is_weekend = calendar.weekday(year, month, day) >= 5

        if is_weekend:
            weekend_days += 1
            weekend_hours += total_hours
        else:
            workday_days += 1
            if total_hours > 8:
                weekday_ot_hours += round_half_down(total_hours - 8)

    # 判断该员工主要班次
    shift_type = '夜班' if night_shift_count > day_shift_count else '白班'

    # 工资计算
    base_salary = round(2100 / 22 * workday_days, 2)
    weekday_ot_salary = round(weekday_ot_hours * 17.9, 2)
    weekend_salary = round(weekend_hours * 23.86, 2)
    night_snack = night_shift_count * 15  # 夜班补贴15元/天
    total_salary = round(base_salary + weekday_ot_salary + weekend_salary + night_snack, 2)

    return {
        'name': name,
        'shift_type': shift_type,
        'workday_days': workday_days,
        'weekend_days': weekend_days,
        'weekday_ot_hours': weekday_ot_hours,
        'weekday_ot_salary': weekday_ot_salary,
        'weekend_hours': weekend_hours,
        'weekend_salary': weekend_salary,
        'base_salary': base_salary,
        'night_shift_days': night_shift_count,
        'night_snack': night_snack,
        'total_salary': total_salary,
    }


def calc_production2_employee(name, records, year, month):
    """
    计算生产部2（计时制）单个员工工资
    规则：
    - 每日工时0.5H阶梯向下取整（<0.5不计，0.5~1计0.5，1~1.5计1...）
    - 当月总工时 × 21元/h
    - 不区分工作日/周末
    """
    total_hours = 0
    work_days = 0

    for rec in records:
        punch_str = rec.get('punch_str', '')
        times = parse_punch_records(punch_str)

        if len(times) < 2:
            continue

        shift = detect_shift(times)
        hours = calc_work_hours(times, shift)

        if hours >= 0.5:
            total_hours += hours
            work_days += 1

    total_salary = round(total_hours * 21, 2)

    return {
        'name': name,
        'work_days': work_days,
        'total_hours': total_hours,
        'hourly_rate': 21,
        'total_salary': total_salary,
    }


def count_sundays(year, month):
    """计算指定月份的周日天数"""
    _, days = calendar.monthrange(year, month)
    return sum(1 for d in range(1, days + 1) if calendar.weekday(year, month, d) == 6)


def calc_tech_employee(name, records, year, month, holidays=0):
    """
    技术部工资计算
    - 应出勤：TECH_SUNDAY_REST人员=当月天数-周日天数，其他=28天
    - 实际出勤：有打卡记录即算出勤（>=1次打卡），无需算加班
    - 基本工资 = 底薪 / 应出勤 × 实际出勤
    - 岗位工资 = 岗位 / 应出勤 × 实际出勤
    - 节假日工资 = (底薪+岗位) / 应出勤 / 12 * 8 * 节假日天数（不计入出勤）
    - 高温补贴：7-10月 150元（仅技术部本部，不含特殊人员）
    """
    if name in TECH_SUNDAY_REST:
        total_days = calendar.monthrange(year, month)[1]
        sundays = count_sundays(year, month)
        required_days = total_days - sundays
    else:
        required_days = 28

    actual_days = 0
    for rec in records:
        punch_str = rec.get('punch_str', '')
        times = parse_punch_records(punch_str)
        if len(times) >= 1:
            actual_days += 1

    salary_info = TECH_SALARY.get(name, {'base': 0, 'position': 0})
    base_total = salary_info['base']
    position_total = salary_info['position']

    if required_days > 0:
        base_salary = round(base_total / required_days * actual_days, 2)
        position_salary = round(position_total / required_days * actual_days, 2)
    else:
        base_salary = 0
        position_salary = 0

    # 节假日工资 = (底薪+岗位) / 应出勤 / 12 * 8 * 节假日天数
    holiday_salary = 0
    if holidays > 0 and required_days > 0:
        holiday_salary = round((base_total + position_total) / required_days / 12 * 8 * holidays, 2)

    # 高温补贴：7-10月，仅技术部本部（不含特殊人员）
    is_tech_native = name not in TECH_SPECIAL
    high_temp = 150 if (month in [7, 8, 9, 10] and is_tech_native) else 0

    total_salary = round(base_salary + position_salary + holiday_salary + high_temp, 2)

    return {
        'name': name,
        'special': name in TECH_SPECIAL,
        'required_days': required_days,
        'actual_days': actual_days,
        'base_total': base_total,
        'position_total': position_total,
        'base_salary': base_salary,
        'position_salary': position_salary,
        'holidays': holidays,
        'holiday_salary': holiday_salary,
        'high_temp': high_temp,
        'total_salary': total_salary,
    }


def calc_ouyang(name, records, year, month, holidays=0):
    """
    欧阳宇专属规则
    - 应出勤 = 当月天数 - 周日天数
    - 只有1个打卡也视为出勤
    - 基本工资2730 / 应出勤 × 实际出勤
    - 岗位工资3970 / 应出勤 × 实际出勤
    - 节假日工资 = (2730+3970) / 应出勤 / 12 * 8 * 节假日天数（不计入出勤）
    - 补贴 = 300 / 应出勤 × 实际出勤
    - 无其他奖金补贴
    """
    total_days = calendar.monthrange(year, month)[1]
    sundays = count_sundays(year, month)
    required_days = total_days - sundays

    actual_days = 0
    for rec in records:
        punch_str = rec.get('punch_str', '')
        times = parse_punch_records(punch_str)
        if len(times) >= 1:  # 1次打卡也算出勤
            actual_days += 1

    base_total = 2730
    position_total = 3970
    subsidy_total = 300

    if required_days > 0:
        base_salary = round(base_total / required_days * actual_days, 2)
        position_salary = round(position_total / required_days * actual_days, 2)
        subsidy = round(subsidy_total / required_days * actual_days, 2)
    else:
        base_salary = 0
        position_salary = 0
        subsidy = 0

    # 节假日工资 = (底薪+岗位) / 应出勤 / 12 * 8 * 节假日天数
    holiday_salary = 0
    if holidays > 0 and required_days > 0:
        holiday_salary = round((base_total + position_total) / required_days / 12 * 8 * holidays, 2)

    total_salary = round(base_salary + position_salary + subsidy + holiday_salary, 2)

    return {
        'name': name,
        'special': True,
        'required_days': required_days,
        'actual_days': actual_days,
        'base_total': base_total,
        'position_total': position_total,
        'base_salary': base_salary,
        'position_salary': position_salary,
        'subsidy': subsidy,
        'holidays': holidays,
        'holiday_salary': holiday_salary,
        'total_salary': total_salary,
    }


def parse_excel(filepath, year, month):
    """解析钉钉考勤报表，提取人员、部门、打卡记录"""
    wb = openpyxl.load_workbook(filepath, read_only=False)
    if '打卡时间' in wb.sheetnames:
        ws = wb['打卡时间']
    else:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # 标题行解析日期范围
    title = str(rows[0][0]) if rows[0][0] else ''
    date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})\s*至\s*(\d{4})-(\d{2})-(\d{2})', title)
    if date_match:
        start_day = int(date_match.group(3))
        end_day = int(date_match.group(6))
        total_days = end_day - start_day + 1
        first_day = start_day
        date_range = f"{year}年{month}月{start_day}日~{end_day}日"
    else:
        total_days = calendar.monthrange(year, month)[1]
        first_day = 1
        date_range = f"{year}年{month}月1日~{total_days}日"

    data_col_start = 6
    employees = []
    seen = set()

    for row in rows[4:]:
        vals = list(row)
        name = vals[0]
        dept = vals[2]

        if not name or dept is None:
            continue

        name = str(name).strip()
        dept_str = str(dept).strip() if dept else ''
        dept_str = dept_str.split('\n')[0].strip()

        if name in seen:
            continue
        seen.add(name)

        has_resigned = '离职' in name
        cname = clean_name(name)

        # 解析每日打卡记录
        records = []
        for day_idx in range(total_days):
            col_idx = data_col_start + day_idx
            day = first_day + day_idx
            punch_raw = vals[col_idx] if col_idx < len(vals) else None
            if punch_raw is not None and not isinstance(punch_raw, str):
                punch_raw = None
            punch_str = ''
            if punch_raw:
                punch_str = punch_raw.replace('\n', '  ').strip()
            records.append({
                'day': day,
                'punch_str': punch_str,
            })

        employees.append({
            'name': cname,
            'raw_name': name,
            'department': dept_str,
            'has_resigned': has_resigned,
            'records': records,
        })

    return {
        'employees': employees,
        'date_range': date_range,
    }


@app.route('/')
def index():
    now = datetime.now()
    return render_template('index.html', current_year=now.year, current_month=now.month)


@app.route('/parse', methods=['POST'])
def parse():
    """解析考勤表，返回人员部门分组 + 生产部工资计算"""
    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        return jsonify({'error': '请上传 .xlsx 格式的考勤表'}), 400

    year = int(request.form.get('year', datetime.now().year))
    month = int(request.form.get('month', datetime.now().month))
    holidays = int(request.form.get('holidays', 0))

    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)

    try:
        data = parse_excel(filepath, year, month)

        # 按部门分组
        dept_groups = {}
        resigned = []
        production_results = []
        production2_results = []
        mold_results = []
        tech_results = []
        # 收集需要走技术部规则的特殊人员（跨部门）
        tech_special_emps = []

        for emp in data['employees']:
            if emp['has_resigned']:
                resigned.append(emp)
                continue

            dept = emp['department'] or '未分配部门'
            dept_groups.setdefault(dept, []).append({
                'name': emp['name'],
                'department': emp['department'],
            })

            # 特殊人员：走技术部规则（来自其他部门）
            if emp['name'] in TECH_SPECIAL and dept != '技术部':
                tech_special_emps.append(emp)
                continue  # 不再走各自部门的计算

            # 欧阳宇：专属规则，归入生产部展示
            if emp['name'] == '欧阳宇':
                result = calc_ouyang(
                    emp['name'], emp['records'], year, month, holidays
                )
                production_results.append(result)
                continue

            # 生产部员工（排除4人）计算工资
            if dept == '生产部' and emp['name'] not in PRODUCTION_EXCLUDE:
                result = calc_production_employee(
                    emp['name'], emp['records'], year, month
                )
                production_results.append(result)

            # 生产部2（计时制）
            if dept == '生产部2':
                result = calc_production2_employee(
                    emp['name'], emp['records'], year, month
                )
                production2_results.append(result)

            # 模房
            if dept == '模房' and emp['name'] not in MOLD_EXCLUDE:
                if emp['name'] == '张翱':
                    result = calc_production_employee(
                        emp['name'], emp['records'], year, month
                    )
                    result['special'] = True
                    mold_results.append(result)
                else:
                    result = calc_production2_employee(
                        emp['name'], emp['records'], year, month
                    )
                    result['special'] = False
                    mold_results.append(result)

            # 技术部本部
            if dept == '技术部':
                result = calc_tech_employee(
                    emp['name'], emp['records'], year, month, holidays
                )
                tech_results.append(result)

        # 计算特殊人员（走技术部规则）
        for emp in tech_special_emps:
            result = calc_tech_employee(
                emp['name'], emp['records'], year, month, holidays
            )
            tech_results.append(result)

        return jsonify({
            'success': True,
            'date_range': data['date_range'],
            'total_count': len(data['employees']),
            'active_count': len(data['employees']) - len(resigned),
            'resigned_count': len(resigned),
            'departments': {dept: emps for dept, emps in sorted(dept_groups.items())},
            'resigned': [{'name': e['name'], 'department': e['department']} for e in resigned],
            'production_results': production_results,
            'production2_results': production2_results,
            'mold_results': mold_results,
            'tech_results': tech_results,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'解析出错: {str(e)}'}), 500
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)


if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    app.run(debug=True, host='0.0.0.0', port=5001)
